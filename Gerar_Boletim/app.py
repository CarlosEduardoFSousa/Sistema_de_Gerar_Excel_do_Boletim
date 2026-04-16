from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
import os

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_EXCEL = os.path.join(BASE_DIR, 'Boletim.xlsx')

def processar_dados_para_formulas(dados_do_form):
    lista_final = []
    materias = dados_do_form.getlist('nome_materia')

    for i, nome in enumerate(materias):
        linha = i + 2
        n1 = dados_do_form.get(f'nota1_{i}', 0) or 0
        n2 = dados_do_form.get(f'nota2_{i}', 0) or 0
        n3 = dados_do_form.get(f'nota3_{i}', 0) or 0
        n4 = dados_do_form.get(f'nota4_{i}', 0) or 0

        f_media = f"=(B{linha}+C{linha}+D{linha}+E{linha})/4"
        f_status = f'=IF(F{linha}>=6, "Aprovado", "Reprovado")'
        f_falta = f"=MAX(0, 24-(B{linha}+C{linha}+D{linha}+E{linha}))"

        lista_final.append({
            'Matéria': nome,
            'B1': float(n1),
            'B2': float(n2),
            'B3': float(n3),
            'B4': float(n4),
            'Média Final': f_media,
            'Status': f_status,
            'Pontos Faltantes (Meta 24)': f_falta
        })
    return lista_final

def gerar_excel_profissional(dados):
    df = pd.DataFrame(dados)
    df.to_excel(OUTPUT_EXCEL, index=False, engine='openpyxl')

    wb = load_workbook(OUTPUT_EXCEL)
    ws = wb.active

    azul_escuro = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    fonte_branca = Font(color="FFFFFF", bold=True)
    centralizar = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = centralizar
            if cell.row == 1:
                cell.fill = azul_escuro
                cell.font = fonte_branca

    verde_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    vermelho_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    regra_aprovado = CellIsRule(operator='equal', formula=['"Aprovado"'], fill=verde_fill)
    regra_reprovado = CellIsRule(operator='equal', formula=['"Reprovado"'], fill=vermelho_fill)

    ws.conditional_formatting.add(f'G2:G{ws.max_row}', regra_aprovado)
    ws.conditional_formatting.add(f'G2:G{ws.max_row}', regra_reprovado)

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 4

    wb.save(OUTPUT_EXCEL)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/gerar', methods=['POST'])
def gerar():
    dados = processar_dados_para_formulas(request.form)
    if not dados:
        return "Adicione pelo menos uma matéria!", 400
    try:
        gerar_excel_profissional(dados)
    except PermissionError:
        return "Erro: O arquivo 'Boletim.xlsx' está aberto. Feche-o e tente novamente.", 403
    return send_file(OUTPUT_EXCEL, as_attachment=True)

@app.route('/importar', methods=['POST'])
def importar():
    if 'arquivo' not in request.files:
        return jsonify({'erro': 'Nenhum arquivo enviado'}), 400

    arquivo = request.files['arquivo']
    if not arquivo.filename.endswith('.xlsx'):
        return jsonify({'erro': 'Envie um arquivo .xlsx válido'}), 400

    try:
        df = pd.read_excel(arquivo, engine='openpyxl')
        colunas = [str(c).strip() for c in df.columns.tolist()]
        col_materia = colunas[0]

        mapa_notas = {}
        for col in colunas:
            col_lower = col.lower()
            if col_lower in ['b1', 'bimestre 1', 'bim 1', '1º bimestre', 'nota1']:
                mapa_notas['B1'] = col
            elif col_lower in ['b2', 'bimestre 2', 'bim 2', '2º bimestre', 'nota2']:
                mapa_notas['B2'] = col
            elif col_lower in ['b3', 'bimestre 3', 'bim 3', '3º bimestre', 'nota3']:
                mapa_notas['B3'] = col
            elif col_lower in ['b4', 'bimestre 4', 'bim 4', '4º bimestre', 'nota4']:
                mapa_notas['B4'] = col

        if len(mapa_notas) < 4:
            notas_keys = ['B1', 'B2', 'B3', 'B4']
            notas_encontradas = []
            for col in colunas[1:]:
                try:
                    if pd.api.types.is_numeric_dtype(df[col]):
                        notas_encontradas.append(col)
                except:
                    pass
                if len(notas_encontradas) == 4:
                    break
            for k, v in zip(notas_keys, notas_encontradas):
                if k not in mapa_notas:
                    mapa_notas[k] = v

        materias = []
        for _, row in df.iterrows():
            nome = str(row[col_materia]) if pd.notna(row[col_materia]) else ''
            if not nome or nome.lower() in ['nan', '']:
                continue

            def pegar_nota(key):
                if key in mapa_notas and mapa_notas[key] in row.index:
                    val = row[mapa_notas[key]]
                    if pd.notna(val):
                        try:
                            return float(val)
                        except:
                            pass
                return 0.0

            materias.append({
                'nome': nome,
                'n1': pegar_nota('B1'),
                'n2': pegar_nota('B2'),
                'n3': pegar_nota('B3'),
                'n4': pegar_nota('B4'),
            })

        return jsonify({'materias': materias, 'total': len(materias)})

    except Exception as e:
        return jsonify({'erro': f'Erro ao ler o arquivo: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True)
